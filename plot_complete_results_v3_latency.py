import csv
import os
from collections import defaultdict

import matplotlib.pyplot as plt
from openpyxl import Workbook

# =========================================================
# CONFIGURACAO V3 - COM LATENCIA EXPLICITA
# =========================================================

BASE_DIR = "v3_dense_sweep_latency"
RESULTS_DIR = os.path.join(BASE_DIR, "results")
FIGURES_DIR = os.path.join(BASE_DIR, "figures")

INPUT_CSV = os.path.join(RESULTS_DIR, "aggregated_results_v3_latency.csv")
EXCEL_OUTPUT = os.path.join(RESULTS_DIR, "results_summary_v3_latency.xlsx")

CSV_DELIMITER = ";"

os.makedirs(FIGURES_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)


# =========================================================
# UTILITARIOS
# =========================================================

def safe_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_int(value):
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def mean(values):
    valid = [v for v in values if v is not None]
    if not valid:
        return None
    return sum(valid) / len(valid)


def unique_sorted(rows, key):
    return sorted({row[key] for row in rows})


def load_aggregated_results(filename):
    rows = []
    with open(filename, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=CSV_DELIMITER)
        for row in reader:
            rows.append({
                "scenario_id": row["scenario_id"],
                "E_K_target": safe_int(row["E_K_target"]),
                "beta_blockage": safe_float(row["beta_blockage"]),
                "blocklength_n": safe_int(row["blocklength_n"]),
                "beam_misalignment_deg": safe_int(row["beam_misalignment_deg"]),
                "num_samples": safe_int(row["num_samples"]),
                "mean_service_distance_m": safe_float(row["mean_service_distance_m"]),
                "mean_K_active": safe_float(row["mean_K_active"]),
                "mean_sinr_db": safe_float(row["mean_sinr_db"]),
                "p5_sinr_db": safe_float(row["p5_sinr_db"]),
                "p10_sinr_db": safe_float(row["p10_sinr_db"]),
                "p50_sinr_db": safe_float(row["p50_sinr_db"]),
                "p90_sinr_db": safe_float(row["p90_sinr_db"]),
                "p95_sinr_db": safe_float(row["p95_sinr_db"]),
                "mean_epsilon_fbl": safe_float(row["mean_epsilon_fbl"]),
                "p50_epsilon_fbl": safe_float(row["p50_epsilon_fbl"]),
                "p90_epsilon_fbl": safe_float(row["p90_epsilon_fbl"]),
                "p95_epsilon_fbl": safe_float(row["p95_epsilon_fbl"]),
                "mean_latency_s": safe_float(row["mean_latency_s"]),
                "p50_latency_s": safe_float(row["p50_latency_s"]),
                "p90_latency_s": safe_float(row["p90_latency_s"]),
                "p95_latency_s": safe_float(row["p95_latency_s"]),
                "reliability_only_1e5": safe_float(row["reliability_only_1e5"]),
                "failure_rate_only_1e5": safe_float(row["failure_rate_only_1e5"]),
                "urllc_success_1ms_1e5": safe_float(row["urllc_success_1ms_1e5"]),
                "urllc_failure_1ms_1e5": safe_float(row["urllc_failure_1ms_1e5"]),
            })
    return rows


def aggregate_for_curve(rows, x_key, y_key, fixed_filters):
    grouped = defaultdict(list)
    for row in rows:
        ok = True
        for k, v in fixed_filters.items():
            if row[k] != v:
                ok = False
                break
        if ok:
            grouped[row[x_key]].append(row[y_key])
    x_vals = sorted(grouped.keys())
    y_vals = [mean(grouped[x]) for x in x_vals]
    return x_vals, y_vals


# =========================================================
# PLOTS
# =========================================================

def save_multi_curve_plot(curves, xlabel, ylabel, title, output_path):
    plt.figure(figsize=(8.8, 5.6))
    for curve in curves:
        plt.plot(curve["x"], curve["y"], marker="o", label=curve["label"])
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close()


def save_heatmap(matrix, x_labels, y_labels, xlabel, ylabel, title, output_path, colorbar_label):
    plt.figure(figsize=(8.5, 6.0))
    plt.imshow(matrix, aspect="auto")
    plt.colorbar(label=colorbar_label)
    plt.xticks(range(len(x_labels)), x_labels)
    plt.yticks(range(len(y_labels)), y_labels)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close()


def build_plot_1(rows):
    curves = []
    for n in unique_sorted(rows, "blocklength_n"):
        x, y = aggregate_for_curve(rows, "E_K_target", "urllc_success_1ms_1e5", {"blocklength_n": n})
        curves.append({"label": f"n={n}", "x": x, "y": y})
    save_multi_curve_plot(curves,
                          xlabel="Mean Number of Active Interferers E[K]",
                          ylabel="Empirical URLLC Success",
                          title="URLLC Success vs E[K] for Different Blocklengths",
                          output_path=os.path.join(FIGURES_DIR, "fig1_urllc_success_vs_EK_by_n_v3.png"))
    return curves


def build_plot_2(rows):
    curves = []
    for beta in unique_sorted(rows, "beta_blockage"):
        x, y = aggregate_for_curve(rows, "blocklength_n", "mean_latency_s", {"beta_blockage": beta})
        curves.append({"label": f"beta={beta}", "x": x, "y": y})
    save_multi_curve_plot(curves,
                          xlabel="Blocklength n",
                          ylabel="Mean Latency (s)",
                          title="Mean Latency vs Blocklength for Different Blockage Regimes",
                          output_path=os.path.join(FIGURES_DIR, "fig2_mean_latency_vs_n_by_beta_v3.png"))
    return curves


def build_plot_3(rows):
    curves = []
    for mis in unique_sorted(rows, "beam_misalignment_deg"):
        x, y = aggregate_for_curve(rows, "E_K_target", "p95_latency_s", {"beam_misalignment_deg": mis})
        curves.append({"label": f"mis={mis} deg", "x": x, "y": y})
    save_multi_curve_plot(curves,
                          xlabel="Mean Number of Active Interferers E[K]",
                          ylabel="95th Percentile Latency (s)",
                          title="95th Percentile Latency vs E[K] for Different Misalignments",
                          output_path=os.path.join(FIGURES_DIR, "fig3_p95_latency_vs_EK_by_misalignment_v3.png"))
    return curves


def build_plot_4(rows):
    curves = []
    for mis in unique_sorted(rows, "beam_misalignment_deg"):
        x, y = aggregate_for_curve(rows, "E_K_target", "p10_sinr_db", {"beam_misalignment_deg": mis})
        curves.append({"label": f"mis={mis} deg", "x": x, "y": y})
    save_multi_curve_plot(curves,
                          xlabel="Mean Number of Active Interferers E[K]",
                          ylabel="10th Percentile SINR (dB)",
                          title="10th Percentile SINR vs E[K] for Different Misalignments",
                          output_path=os.path.join(FIGURES_DIR, "fig4_p10_sinr_vs_EK_by_misalignment_v3.png"))
    return curves


def build_plot_5(rows):
    curves = []
    for beta in unique_sorted(rows, "beta_blockage"):
        x, y = aggregate_for_curve(rows, "blocklength_n", "urllc_success_1ms_1e5", {"beta_blockage": beta})
        curves.append({"label": f"beta={beta}", "x": x, "y": y})
    save_multi_curve_plot(curves,
                          xlabel="Blocklength n",
                          ylabel="Empirical URLLC Success",
                          title="URLLC Success vs Blocklength for Different Blockage Regimes",
                          output_path=os.path.join(FIGURES_DIR, "fig5_urllc_success_vs_n_by_beta_v3.png"))
    return curves


def build_heatmap_1(rows):
    n_fixed = max(unique_sorted(rows, "blocklength_n"))
    mis_fixed = 0
    ek_values = unique_sorted(rows, "E_K_target")
    beta_values = unique_sorted(rows, "beta_blockage")

    matrix = []
    table_rows = []
    for beta in beta_values:
        row_vals = []
        for ek in ek_values:
            vals = [
                r["urllc_success_1ms_1e5"]
                for r in rows
                if r["E_K_target"] == ek and r["beta_blockage"] == beta and r["blocklength_n"] == n_fixed and r["beam_misalignment_deg"] == mis_fixed
            ]
            m = mean(vals)
            row_vals.append(m if m is not None else 0.0)
            table_rows.append({
                "beta_blockage": beta,
                "E_K_target": ek,
                "blocklength_n_fixed": n_fixed,
                "beam_misalignment_deg_fixed": mis_fixed,
                "urllc_success_1ms_1e5": m
            })
        matrix.append(row_vals)

    save_heatmap(matrix,
                 x_labels=ek_values,
                 y_labels=beta_values,
                 xlabel="E[K]",
                 ylabel="beta",
                 title=f"URLLC Success Heatmap (n={n_fixed}, misalignment={mis_fixed} deg)",
                 output_path=os.path.join(FIGURES_DIR, "fig6_heatmap_urllc_success_EK_beta_v3.png"),
                 colorbar_label="URLLC Success")
    return table_rows


def build_heatmap_2(rows):
    ek_values = unique_sorted(rows, "E_K_target")
    beta_values = unique_sorted(rows, "beta_blockage")
    ek_fixed = ek_values[len(ek_values) // 2]
    beta_fixed = beta_values[len(beta_values) // 2]
    n_values = unique_sorted(rows, "blocklength_n")
    mis_values = unique_sorted(rows, "beam_misalignment_deg")

    matrix = []
    table_rows = []
    for mis in mis_values:
        row_vals = []
        for n in n_values:
            vals = [
                r["mean_latency_s"]
                for r in rows
                if r["E_K_target"] == ek_fixed and r["beta_blockage"] == beta_fixed and r["blocklength_n"] == n and r["beam_misalignment_deg"] == mis
            ]
            m = mean(vals)
            row_vals.append(m if m is not None else 0.0)
            table_rows.append({
                "E_K_target_fixed": ek_fixed,
                "beta_blockage_fixed": beta_fixed,
                "blocklength_n": n,
                "beam_misalignment_deg": mis,
                "mean_latency_s": m
            })
        matrix.append(row_vals)

    save_heatmap(matrix,
                 x_labels=n_values,
                 y_labels=mis_values,
                 xlabel="Blocklength n",
                 ylabel="Misalignment (deg)",
                 title=f"Mean Latency Heatmap (E[K]={ek_fixed}, beta={beta_fixed})",
                 output_path=os.path.join(FIGURES_DIR, "fig7_heatmap_mean_latency_n_misalignment_v3.png"),
                 colorbar_label="Mean Latency (s)")
    return table_rows


# =========================================================
# EXCEL
# =========================================================

def auto_adjust_width(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)


def write_sheet_from_dicts(ws, rows):
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    auto_adjust_width(ws)


def write_curve_sheet(ws, x_name, curves):
    headers = [x_name] + [curve["label"] for curve in curves]
    ws.append(headers)
    all_x = sorted(set(x for curve in curves for x in curve["x"]))
    for xv in all_x:
        row = [xv]
        for curve in curves:
            mapping = dict(zip(curve["x"], curve["y"]))
            row.append(mapping.get(xv, None))
        ws.append(row)
    auto_adjust_width(ws)


def export_to_excel(all_rows, plot1, plot2, plot3, plot4, plot5, heat1, heat2):
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "raw_aggregated_v3"
    write_sheet_from_dicts(ws0, all_rows)

    ws1 = wb.create_sheet("urllc_vs_EK_by_n_v3")
    write_curve_sheet(ws1, "E_K_target", plot1)

    ws2 = wb.create_sheet("latency_vs_n_by_beta_v3")
    write_curve_sheet(ws2, "blocklength_n", plot2)

    ws3 = wb.create_sheet("p95lat_vs_EK_by_mis_v3")
    write_curve_sheet(ws3, "E_K_target", plot3)

    ws4 = wb.create_sheet("p10sinr_vs_EK_by_mis_v3")
    write_curve_sheet(ws4, "E_K_target", plot4)

    ws5 = wb.create_sheet("urllc_vs_n_by_beta_v3")
    write_curve_sheet(ws5, "blocklength_n", plot5)

    ws6 = wb.create_sheet("heatmap_urllc_EK_beta_v3")
    write_sheet_from_dicts(ws6, heat1)

    ws7 = wb.create_sheet("heatmap_latency_n_mis_v3")
    write_sheet_from_dicts(ws7, heat2)

    wb.save(EXCEL_OUTPUT)


# =========================================================
# MAIN
# =========================================================

def main():
    print("==============================================")
    print(" Plot Complete Results v3 - Dense Sweep + Latency")
    print("==============================================")
    print(f"Input CSV : {INPUT_CSV}")
    print(f"Figures   : {FIGURES_DIR}")
    print(f"Excel     : {EXCEL_OUTPUT}")
    print("==============================================")

    rows = load_aggregated_results(INPUT_CSV)
    plot1 = build_plot_1(rows)
    plot2 = build_plot_2(rows)
    plot3 = build_plot_3(rows)
    plot4 = build_plot_4(rows)
    plot5 = build_plot_5(rows)
    heat1 = build_heatmap_1(rows)
    heat2 = build_heatmap_2(rows)

    export_to_excel(rows, plot1, plot2, plot3, plot4, plot5, heat1, heat2)

    print("==============================================")
    print("Processamento v3 concluido com sucesso.")
    print(f"Figuras guardadas em: {FIGURES_DIR}")
    print(f"Excel guardado em   : {EXCEL_OUTPUT}")
    print("==============================================")


if __name__ == "__main__":
    main()
